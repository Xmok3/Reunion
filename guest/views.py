import os
import csv
from datetime import datetime
from django.conf import settings
from django.http import HttpResponse
from django.contrib import messages
import qrcode
from PIL import Image
from io import BytesIO, StringIO
from django.core.files.base import ContentFile
from django.shortcuts import render, redirect
from .models import Guest
from django.core.mail import EmailMessage
from django.conf import settings
from io import BytesIO

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

def register(request):
    if request.method == "POST":
        full_name = request.POST.get("full_name")
        phone_number = request.POST.get("phone_number")
        email = request.POST.get("email")

        # Save guest
        guest = Guest.objects.create(
            full_name=full_name,
            phone_number=phone_number,
            email=email
        )

        # Generate QR code
        qr = qrcode.QRCode(error_correction=qrcode.constants.ERROR_CORRECT_H)
        qr.add_data(guest.qr_code_value)
        qr.make(fit=True)
        qr_img = qr.make_image(fill_color="black", back_color="white").convert("RGBA")

        # Open event logo
        logo_path = os.path.join(settings.BASE_DIR, "guest", "static", "images", "event-logo.jpg")
        logo = Image.open(logo_path).convert("RGBA")
        logo_width, logo_height = logo.size

        # Resize QR to fit nicely
        qr_width, qr_height = qr_img.size
        factor = 4  # smaller QR
        qr_img = qr_img.resize((qr_width//factor, qr_height//factor))

        # Paste QR at bottom left
        position = (10, logo_height - qr_img.height - 10)
        logo.paste(qr_img, position, qr_img)

        # Save final image
        buffer = BytesIO()
        logo.save(buffer, format="PNG")
        guest.qr_image.save(f"{guest.qr_code_value}.png", ContentFile(buffer.getvalue()))
        
        # --- SEND EMAIL WITH QR ---

        subject = "Your Event QR Code"

        message = f"""
        Hello {guest.full_name},

        Thank you for registering.

        Your QR code is attached to this email.
        Please keep it safe as it will be required for check-in.

        Regards,
        XMOK3 Developments
        """

        email = EmailMessage(
        subject,
        message,
        settings.EMAIL_HOST_USER,
        [guest.email]
        )

        email.attach_file(guest.qr_image.path)
        email.send()


        return redirect("success", code=guest.qr_code_value)

    return render(request, "register.html")



def success(request, code):
    guest = Guest.objects.get(qr_code_value=code)
    return render(request, "success.html", {"guest": guest})


def dashboard(request):
    guests = Guest.objects.all().order_by('-id')
    total_guests = guests.count()
    context = {
        'guests': guests,
        'total_guests': total_guests,
    }
    return render(request, "dashboard.html", context)


def export_csv(request):
    """Export all guests to CSV format"""
    guests = Guest.objects.all().order_by('-id')
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="guests_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Full Name', 'Email', 'Phone Number', 'QR Code'])
    
    for guest in guests:
        writer.writerow([guest.full_name, guest.email, guest.phone_number, guest.qr_code_value])
    
    return response


def export_xlsx(request):
    """Export all guests to XLSX format"""
    if not HAS_OPENPYXL:
        messages.error(request, 'openpyxl is not installed. Please install it to use XLSX export.')
        return redirect('dashboard')
    
    guests = Guest.objects.all().order_by('-id')
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Guests"
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    
    # Create header row
    headers = ['Full Name', 'Email', 'Phone Number', 'QR Code']
    ws.append(headers)
    
    # Style header row
    header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add guest data
    for idx, guest in enumerate(guests, start=2):
        ws.append([guest.full_name, guest.email, guest.phone_number, guest.qr_code_value])
        
        # Alternate row coloring
        if idx % 2 == 0:
            fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
            for cell in ws[idx]:
                cell.fill = fill
    
    # Set response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="guests_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response)
    
    return response


def import_guests(request):
    """Import guests from CSV or XLSX file"""
    if request.method == 'POST':
        file = request.FILES.get('file')
        
        if not file:
            messages.error(request, 'Please select a file to import.')
            return redirect('dashboard')
        
        imported_count = 0
        skipped_count = 0
        errors = []
        
        try:
            if file.name.endswith('.csv'):
                # Handle CSV - try different encodings
                try:
                    decoded_file = file.read().decode('utf-8').splitlines()
                except:
                    file.seek(0)
                    decoded_file = file.read().decode('latin-1').splitlines()
                
                reader = csv.DictReader(decoded_file)
                
                if not reader.fieldnames:
                    messages.error(request, 'CSV file appears to be empty or invalid format.')
                    return redirect('dashboard')
                
                for row_num, row in enumerate(reader, start=2):
                    try:
                        if not row or all(v is None or str(v).strip() == '' for v in row.values()):
                            continue
                            
                        full_name = str(row.get('Full Name', '') or '').strip()
                        email = str(row.get('Email', '') or '').strip()
                        phone_number = str(row.get('Phone Number', '') or '').strip()
                        
                        if not full_name or not email or not phone_number:
                            skipped_count += 1
                            errors.append(f"Row {row_num}: Missing required fields")
                            continue
                        
                        # Check if guest already exists
                        if Guest.objects.filter(email=email).exists():
                            skipped_count += 1
                            errors.append(f"Row {row_num}: Email {email} already exists")
                            continue
                        
                        # Create guest
                        guest = Guest.objects.create(
                            full_name=full_name,
                            email=email,
                            phone_number=phone_number
                        )
                        
                        # Generate QR code
                        try:
                            qr = qrcode.QRCode(error_correction=qrcode.constants.ERROR_CORRECT_H)
                            qr.add_data(guest.qr_code_value)
                            qr.make(fit=True)
                            qr_img = qr.make_image(fill_color="black", back_color="white").convert("RGBA")

                            logo_path = os.path.join(settings.BASE_DIR, "guest", "static", "images", "event-logo.jpg")
                            logo = Image.open(logo_path).convert("RGBA")
                            logo_width, logo_height = logo.size

                            qr_width, qr_height = qr_img.size
                            factor = 4
                            qr_img = qr_img.resize((qr_width//factor, qr_height//factor))

                            position = (10, logo_height - qr_img.height - 10)
                            logo.paste(qr_img, position, qr_img)

                            buffer = BytesIO()
                            logo.save(buffer, format="PNG")
                            guest.qr_image.save(f"{guest.qr_code_value}.png", ContentFile(buffer.getvalue()))
                        except Exception as qr_error:
                            print(f"QR Code generation error: {qr_error}")
                        
                        imported_count += 1
                        
                    except Exception as e:
                        print(f"Row {row_num} error: {str(e)}")
                        skipped_count += 1
                        errors.append(f"Row {row_num}: {str(e)[:50]}")
            
            elif file.name.endswith(('.xlsx', '.xls')):
                # Handle XLSX
                if not HAS_OPENPYXL:
                    messages.error(request, 'openpyxl is not installed. Cannot import XLSX files.')
                    return redirect('dashboard')
                
                try:
                    wb = openpyxl.load_workbook(file)
                    ws = wb.active
                    
                    if not ws:
                        messages.error(request, 'Excel file appears to be empty.')
                        return redirect('dashboard')
                    
                    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                        try:
                            if not row or all(v is None or str(v).strip() == '' for v in row):
                                continue
                                
                            full_name = str(row[0] or '').strip()
                            email = str(row[1] or '').strip()
                            phone_number = str(row[2] or '').strip()
                            
                            if not full_name or not email or not phone_number:
                                skipped_count += 1
                                errors.append(f"Row {row_num}: Missing required fields")
                                continue
                            
                            # Check if guest already exists
                            if Guest.objects.filter(email=email).exists():
                                skipped_count += 1
                                errors.append(f"Row {row_num}: Email {email} already exists")
                                continue
                            
                            # Create guest
                            guest = Guest.objects.create(
                                full_name=full_name,
                                email=email,
                                phone_number=phone_number
                            )
                            
                            # Generate QR code
                            try:
                                qr = qrcode.QRCode(error_correction=qrcode.constants.ERROR_CORRECT_H)
                                qr.add_data(guest.qr_code_value)
                                qr.make(fit=True)
                                qr_img = qr.make_image(fill_color="black", back_color="white").convert("RGBA")

                                logo_path = os.path.join(settings.BASE_DIR, "guest", "static", "images", "event-logo.jpg")
                                logo = Image.open(logo_path).convert("RGBA")
                                logo_width, logo_height = logo.size

                                qr_width, qr_height = qr_img.size
                                factor = 4
                                qr_img = qr_img.resize((qr_width//factor, qr_height//factor))

                                position = (10, logo_height - qr_img.height - 10)
                                logo.paste(qr_img, position, qr_img)

                                buffer = BytesIO()
                                logo.save(buffer, format="PNG")
                                guest.qr_image.save(f"{guest.qr_code_value}.png", ContentFile(buffer.getvalue()))
                            except Exception as qr_error:
                                print(f"QR Code generation error: {qr_error}")
                            
                            imported_count += 1
                            
                        except Exception as e:
                            print(f"Row {row_num} error: {str(e)}")
                            skipped_count += 1
                            errors.append(f"Row {row_num}: {str(e)[:50]}")
                except Exception as excel_error:
                    messages.error(request, f'Error reading Excel file: {str(excel_error)[:100]}')
                    print(f"Excel error: {str(excel_error)}")
                    return redirect('dashboard')
            
            else:
                messages.error(request, 'Please upload a CSV or XLSX file.')
                return redirect('dashboard')
            
            # Show results
            if imported_count > 0:
                messages.success(request, f'✓ Successfully imported {imported_count} guest(s)!')
            if skipped_count > 0:
                error_msg = f'⚠ Skipped {skipped_count} rows.'
                if errors:
                    error_msg += f' Reasons: ' + '; '.join(errors[:3])
                    if len(errors) > 3:
                        error_msg += f'... and {len(errors) - 3} more'
                messages.warning(request, error_msg)
            
            if imported_count == 0 and skipped_count == 0:
                messages.info(request, 'No data to import from file.')
            
        except Exception as e:
            print(f"Import error: {str(e)}")
            messages.error(request, f'Error importing file: {str(e)[:100]}')
        
        return redirect('dashboard')
    
    return redirect('dashboard')

